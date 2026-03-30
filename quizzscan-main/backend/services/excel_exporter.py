"""
Export Excel au format imposé par la plateforme :
Col A: ID étudiant | Col B: vide | Col C: vide | Col D: OUI | Col E+: Q1, Q2...
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Optional
import os


def export_to_excel(
    output_path: str,
    student_results: List[dict],
    nb_questions: int,
) -> str:
    """
    Génère le fichier Excel final.
    student_results: liste de dicts {student_id, answers: {"1": "ABD", ...}}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résultats"

    # ─── En-têtes ────────────────────────────────────────────────────────────
    headers = [
        "ID UTILISATEUR (au choix avec B1,C1)",
        "CODE IDENTIFIANT (au choix avec A1, C1)",
        "PSEUDO ou EMAIL (au choix avec A1,B1)",
        "Présent ? OUI ou NON",
    ]
    for q in range(1, nb_questions + 1):
        headers.append(f"Q{q}")

    # Style en-têtes
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 30

    # ─── Données étudiants ───────────────────────────────────────────────────
    row_fill_even = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    data_align = Alignment(horizontal="center", vertical="center")

    for row_idx, student in enumerate(student_results, 2):
        fill = row_fill_even if row_idx % 2 == 0 else PatternFill(fill_type=None)

        # Col A : ID (numérique si possible)
        sid = student.get("student_id", "")
        try:
            sid_val = int(sid)
        except (ValueError, TypeError):
            sid_val = sid

        ws.cell(row=row_idx, column=1, value=sid_val)
        ws.cell(row=row_idx, column=2, value=None)   # vide - IMPORTANT
        ws.cell(row=row_idx, column=3, value=None)   # vide - IMPORTANT
        ws.cell(row=row_idx, column=4, value="OUI")

        for q in range(1, nb_questions + 1):
            ans = student.get("answers", {}).get(str(q), "")
            cell = ws.cell(row=row_idx, column=4 + q, value=ans)
            cell.fill = fill
            cell.alignment = data_align
            cell.border = thin_border

        # Style colonnes ID et Présent
        for col in [1, 4]:
            c = ws.cell(row=row_idx, column=col)
            c.fill = fill
            c.alignment = data_align
            c.border = thin_border

    # ─── Largeurs colonnes ───────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    for q in range(1, nb_questions + 1):
        col_letter = get_column_letter(4 + q)
        ws.column_dimensions[col_letter].width = 7

    # ─── Figer la 1ère ligne ─────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
